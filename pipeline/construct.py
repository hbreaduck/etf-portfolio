"""
construct.py — 타깃 포트폴리오 산출 + Excel 출력

배분 알고리즘:
  1. 스코어 상위 target_names 종목 선발 (버킷/섹터 제약 없음)
  2. name_cap 초과 종목을 상한에 고정하고, 잉여를 나머지에 재분배 (waterfall)
  3. name_floor 미만 종목 제거 후 재정규화
  현금 cash_pct(5%) 는 별도 고정
"""

import pathlib
import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 핵심 배분 함수 ─────────────────────────────────────────────────────────────

def _waterfall(scores: dict[str, float], target: float, cap: float) -> dict[str, float]:
    """cap 초과 종목을 순서대로 고정하고 잉여를 나머지에 재분배."""
    remaining = target
    weights:  dict[str, float] = {}
    pool = dict(scores)

    for _ in range(len(pool) + 2):
        if not pool or remaining <= 1e-12:
            break
        total = sum(pool.values())
        trial = {t: (s / total) * remaining for t, s in pool.items()}

        over = {t for t, w in trial.items() if w > cap}
        if not over:
            weights.update(trial)
            break

        for t in over:
            weights[t] = cap
        remaining -= len(over) * cap
        pool = {t: s for t, s in pool.items() if t not in over}

    return weights


def _allocate(scores: dict[str, float],
              target: float,
              cap: float,
              floor: float) -> dict[str, float]:
    """
    waterfall cap → drop floor → renormalize.
    플로어 제거 후 재정규화로 cap이 재위반될 수 있으므로 수렴까지 반복.
    """
    pool = {t: s for t, s in scores.items() if s > 0}

    for _ in range(30):
        w = _waterfall(pool, target, cap)

        # 플로어 미만 제거
        w = {t: v for t, v in w.items() if v >= floor}
        if not w:
            return {}

        # 재정규화
        tot = sum(w.values())
        w = {t: v * target / tot for t, v in w.items()}

        # cap 재위반 없으면 수렴
        if all(v <= cap + 1e-9 for v in w.values()):
            return w

        # 재위반 종목이 있으면 현재 가중치를 점수로 삼아 재시도
        pool = dict(w)

    return w


# ── 메인 함수 ──────────────────────────────────────────────────────────────────

def run(scored: pd.DataFrame, cfg: dict, date_str: str) -> pd.DataFrame:
    """
    scored: score.run() 반환값
    반환: 타깃 포트폴리오 DataFrame
    """
    cap     = float(cfg["name_cap"])
    floor   = float(cfg["name_floor"])
    n_names = int(cfg.get("target_names", 15))
    w_cash  = float(cfg.get("cash_pct", 0.05))
    w_eq    = 1.0 - w_cash

    # 스코어 상위 n_names 종목 선발 (버킷/섹터 제약 없음)
    candidates = scored[scored["bucket"] != "cash"].copy()
    candidates = candidates[candidates["score"] > 0]
    top = candidates.nlargest(n_names, "score")

    scores_dict = dict(zip(top["ticker"], top["score"]))
    weights = _allocate(scores_dict, w_eq, cap, floor)

    # 결합
    score_lookup = dict(zip(scored["ticker"], scored["score"]))
    info_lookup  = scored.set_index("ticker")[["name", "gics_sector",
                                               "tech_economic", "bucket"]].to_dict("index")

    rows = []
    for ticker, weight in sorted(weights.items(), key=lambda x: -x[1]):
        info = info_lookup.get(ticker, {})
        rows.append({
            "ticker":        ticker,
            "name":          info.get("name", ""),
            "gics_sector":   info.get("gics_sector", ""),
            "tech_economic": info.get("tech_economic", ""),
            "bucket":        info.get("bucket", ""),
            "score":         round(score_lookup.get(ticker, 0), 4),
            "target_weight": round(weight, 6),
        })

    rows.append({
        "ticker":        "CASH",
        "name":          "현금",
        "gics_sector":   "",
        "tech_economic": "",
        "bucket":        "cash",
        "score":         0,
        "target_weight": round(w_cash, 6),
    })

    port = pd.DataFrame(rows)
    port["target_pct"] = (port["target_weight"] * 100).round(2)
    # 현금을 마지막으로, 나머지는 비중 내림차순
    port = pd.concat([
        port[port["bucket"] != "cash"].sort_values("target_weight", ascending=False),
        port[port["bucket"] == "cash"],
    ], ignore_index=True)

    # 검증
    total  = port["target_weight"].sum()
    n_eq   = len(port[port["bucket"] != "cash"])
    print(f"  [construct] 총 {n_eq}종목 + 현금 | 합계={total*100:.2f}%")
    by_bkt = port.groupby("bucket")["target_weight"].sum()
    for bkt in ["tech", "other", "cash"]:
        sub = port[port["bucket"] == bkt]
        if not sub.empty:
            print(f"    {bkt:6s}: {len(sub)}종목, {sub['target_weight'].sum()*100:.2f}%")

    # Excel 저장
    out_dir = pathlib.Path(cfg["paths"].get("output_dir", "output"))
    out_dir.mkdir(parents=True, exist_ok=True)
    cap_tag  = f"_cap{int(float(cfg['name_cap'])*100):02d}"
    out_path = out_dir / f"portfolio_{date_str}{cap_tag}.xlsx"
    base, suffix = out_path.stem, out_path.suffix
    seq = 0
    while out_path.exists():
        try:
            out_path.open("r+b").close()
            break
        except PermissionError:
            seq += 1
            out_path = out_dir / f"{base}_{seq}{suffix}"
    _save_excel(port, scored, cfg, str(out_path), date_str)

    return port


# ── Excel 출력 ─────────────────────────────────────────────────────────────────

_BUCKET_COLOR = {"tech": "D6E4F7", "other": "E2EFDA", "cash": "FFF2CC"}
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1F3864")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _h(ws, row, col, val, bold=False, fill=None, align="left", num_fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = _BORDER
    if fill:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill)
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _save_excel(port: pd.DataFrame, scored: pd.DataFrame,
                cfg: dict, path: str, date_str: str):
    from openpyxl import Workbook
    wb = Workbook()

    _sheet_portfolio(wb.active, port, date_str)
    wb.active.title = "타깃포트폴리오"

    ws2 = wb.create_sheet("스코어유니버스")
    _sheet_universe(ws2, scored)

    ws3 = wb.create_sheet("버킷요약")
    _sheet_summary(ws3, port, cfg)

    wb.save(path)
    print(f"  [construct] Excel 저장 → {path}")


def _sheet_portfolio(ws, port: pd.DataFrame, date_str: str):
    # 타이틀
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value = f"타깃 포트폴리오  ({date_str[:4]}-{date_str[4:6]}-{date_str[6:]})"
    tc.font = Font(bold=True, size=13, color="1F3864")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = ["버킷", "티커", "종목명", "GICS 섹터", "Tech경제", "스코어(합산비중)", "타깃비중(%)"]
    col_widths = [8, 14, 38, 26, 8, 16, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 18

    for ri, row in enumerate(port.itertuples(index=False), 3):
        fill_color = _BUCKET_COLOR.get(row.bucket, "FFFFFF")
        bkt_label  = {"tech": "테크", "other": "기타", "cash": "현금"}.get(row.bucket, row.bucket)
        data = [bkt_label, row.ticker, row.name, row.gics_sector,
                row.tech_economic, row.score, row.target_pct]
        fmts = [None, None, None, None, None, "0.00", "0.00"]
        aligns = ["center", "left", "left", "left", "center", "right", "right"]

        for ci, (val, fmt, aln) in enumerate(zip(data, fmts, aligns), 1):
            _h(ws, ri, ci, val, fill=fill_color, align=aln, num_fmt=fmt)

    ws.freeze_panes = "A3"


def _sheet_universe(ws, scored: pd.DataFrame):
    etf_cols = ["456600", "426030", "00015B0", "466950"]
    headers = ["티커", "종목명", "GICS섹터", "버킷", "스코어"] + etf_cols
    col_widths = [14, 38, 26, 12, 10] + [10] * len(etf_cols)
    n_cols = len(headers)

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws["A1"].value = "스코어 유니버스 (stock only)"
    ws["A1"].font = Font(bold=True, size=12, color="1F3864")

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(scored.itertuples(index=False), 3):
        bkt_label = {"tech": "테크", "other": "기타", "cash": "현금",
                     "unclassified": "미분류"}.get(row.bucket, row.bucket)
        vals = [row.ticker, row.name, row.gics_sector, bkt_label, row.score]
        for ec in etf_cols:
            vals.append(getattr(row, ec, float("nan")))

        fill_color = _BUCKET_COLOR.get(row.bucket, "F5F5F5")
        fmts = [None, None, None, None, "0.00"] + ["0.00"] * len(etf_cols)
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            aln = "right" if ci >= 5 else "left"
            _h(ws, ri, ci, v, fill=fill_color, align=aln, num_fmt=fmt)

    ws.freeze_panes = "A3"


def _sheet_summary(ws, port: pd.DataFrame, cfg: dict):
    ws["A1"].value = "섹터 분포 요약"
    ws["A1"].font = Font(bold=True, size=12, color="1F3864")

    headers = ["GICS 섹터", "실현(%)", "종목수"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(ci)].width = 28 if ci == 1 else 12

    equity = port[port["bucket"] != "cash"]
    by_sector = (equity.groupby("gics_sector")
                 .agg(실현=("target_weight", "sum"), 종목수=("ticker", "count"))
                 .reset_index().sort_values("실현", ascending=False))

    ri = 3
    for _, row in by_sector.iterrows():
        fill = "F5F5F5"
        for ci, val in enumerate([row["gics_sector"], row["실현"]*100, int(row["종목수"])], 1):
            fmt = "0.00" if ci == 2 else None
            _h(ws, ri, ci, val, fill=fill, align="right" if ci > 1 else "left", num_fmt=fmt)
        ri += 1

    # 현금 행
    cash_w = port[port["bucket"] == "cash"]["target_weight"].sum()
    for ci, val in enumerate(["현금", cash_w*100, 1], 1):
        fmt = "0.00" if ci == 2 else None
        _h(ws, ri, ci, val, fill=_BUCKET_COLOR["cash"],
           align="right" if ci > 1 else "left", num_fmt=fmt)
